import csv
from operator import itemgetter

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter


def csv_reader():
    """Чтение данных из csv файла
            Returns:
                headings (list): список из заголовков фалйла
                vacancies (list): список из профессий
    """
    headlines_list = vacancies_list = []
    length = rows_count = 0
    first_element = True
    with open(file_name, encoding="utf-8-sig") as File:
        reader = csv.reader(File)
        for row in reader:
            rows_count += 1
            if first_element:
                headlines_list = row
                length = len(row)
                first_element = False
            else:
                flag_to_continue = False
                if length != len(row):
                    flag_to_continue = True
                for word in row:
                    if word == "":
                        flag_to_continue = True
                if flag_to_continue:
                    continue
                vacancies_list.append(row)
    if rows_count == 0:
        print("Пустой файл")
        exit()
    elif rows_count == 1:
        print("Нет данных")
        exit()
    return headlines_list, vacancies_list


def csv_filter(reader, list_naming):
    """Записывает в словари данные полученные после чтения csv файла
            Args:
                reader (list): список вакансий полученных после обработки csv файла
                list_naming (list): список заголовков полученных после обработки csv файла
            Returns:
                dictionaries (list[dict]): список словарей с вакансиями
    """
    dictionaries_list = []
    for vacancy in reader:
        dictionary = {}
        for i in range(len(list_naming)):
            dictionary[list_naming[i]] = vacancy[i]
        dictionaries_list.append(dictionary)
    return dictionaries_list


def dict_sorter(dictionary):
    """Сортирует словарь по ключам
    """
    return dict(sorted(dictionary.items(), key=itemgetter(0)))


def dict_processing_published(dictionary, vacancy, addend):
    """
    Обрабатывает словарь на места публикации вакансиии
    """
    if vacancy.published_at in dictionary:
        dictionary[vacancy.published_at] += addend
    else:
        dictionary[vacancy.published_at] = addend
    return dictionary


def dict_processing_area(dictionary, vacancy, addend):
    """
        Обрабатывает словарь на города вакансий
    """
    if vacancy.area_name in dictionary:
        dictionary[vacancy.area_name] += addend
    else:
        dictionary[vacancy.area_name] = addend
    return dictionary


def take_ten_items(dictionary):
    """
        Функция возвращающая первые десять элемнтов словаря
        Parameters:
                   dictionary (dict): словарь
        Returns:
                  (dict): новый словарь из 10 элементов
    >>> take_ten_items({"a": 1, "b": 3, "c": 4, "d": 0.7, "e": 6, "f": 7,"g": 8, "h": 9, "i": 10, "j": 11, "k": 12, "l": 13.7, "m": 20})
    {'a': 1, 'b': 3, 'c': 4, 'd': 0.7, 'e': 6, 'f': 7, 'g': 8, 'h': 9, 'i': 10, 'j': 11}
    >>> take_ten_items({'a': 1, 'b': 3, 'c': 4, 'd': 0.7, 'e': 6})
    {'a': 1, 'b': 3, 'c': 4, 'd': 0.7, 'e': 6}
    """
    new_dictionary = {}
    i = 0
    for key in dictionary:
        new_dictionary[key] = round(dictionary[key], 4)
        i += 1
        if i == 10:
            break
    return new_dictionary


class DataSet:
    """Считывание файла и формирование структуры данных о нем.

    Attributes:
            file_name (str): название csv файла.
            profession (str): Название профессии.
    """
    def __init__(self, file_name, profession):
        """Инициализирует класс DataSet. Чтение фала,форматирование,вывод информации.
                        Args:
                            file_name (str): название csv файла.
                            profession (str): название профессии.
        """
        self.file_name = file_name
        self.profession = profession
        headlines, vacancies = csv_reader()
        dictionaries = csv_filter(vacancies, headlines)
        self.vacancies_objects = [Vacancy(dictionary) for dictionary in dictionaries]
        self.vacancies_count_by_years = self.get_years_count_vacancies()
        self.vacancies_count_by_years_for_profession = self.get_years_vacancies_count_for_profession()
        self.salary_by_years = self.get_years_salary()
        self.salary_by_years_for_profession = self.get_years_salary_profession()
        self.vacancies_count_by_cities = self.get_count_cities_vacancies()
        self.vacancies_share_by_cities = self.get_share_cities_vacancies()
        self.salary_by_cities = self.get_cities_salary()

    def get_years_count_vacancies(self):
        """Возвращает количество ваканчий за год"""
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary = dict_processing_published(dictionary, vacancy, 1)
        dictionary = dict_sorter(dictionary)
        return dictionary

    def get_years_vacancies_count_for_profession(self):
        """Возвращает количество ваканчий за год по профессиям"""
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            dictionary = dict_processing_published(dictionary, vacancy, 1)
        dictionary = dict_sorter(dictionary)
        if len(dictionary) == 0:
            dictionary = {2022: 0}
        return dictionary

    def get_years_salary(self):
        """Возвращает зарплаты за год"""
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary = dict_processing_published(dictionary, vacancy, vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_years[key])
        dictionary = dict_sorter(dictionary)
        return dictionary

    def get_years_salary_profession(self):
        """Возвращает зарплаты за год по профессиям"""
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            dictionary = dict_processing_published(dictionary, vacancy, vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_years_for_profession[key])
        dictionary = dict_sorter(dictionary)
        if len(dictionary) == 0:
            dictionary = {2022: 0}
        return dictionary

    def get_count_cities_vacancies(self):
        """Возвращает количество вакансий в городах"""
        dictionary = {}
        for vacancy in self.vacancies_objects:
            dictionary = dict_processing_area(dictionary, vacancy, 1)
        return dictionary

    def get_share_cities_vacancies(self):
        dictionary = {}
        for key in self.vacancies_count_by_cities:
            if self.vacancies_count_by_cities[key] / len(self.vacancies_objects) >= 0.01:
                dictionary[key] = self.vacancies_count_by_cities[key] / len(self.vacancies_objects)
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True))
        new_dictionary = take_ten_items(dictionary)
        return new_dictionary

    def get_cities_salary(self):
        """Возвращает зарплату в городах"""
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.vacancies_count_by_cities[vacancy.area_name] / len(self.vacancies_objects) < 0.01:
                continue
            dictionary = dict_processing_area(dictionary, vacancy, vacancy.salary)
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_cities[key])
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True))
        new_dictionary = take_ten_items(dictionary)
        return new_dictionary

    def print_info(self):
        """Выводит данные необходимы для графиков в консоль
        """
        print(f"Динамика уровня зарплат по годам: {str(self.salary_by_years)}")
        print(f"Динамика количества вакансий по годам: {str(self.vacancies_count_by_years)}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {str(self.salary_by_years_for_profession)}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {str(self.vacancies_count_by_years_for_profession)}")
        print(f"Уровень зарплат по городам (в порядке убывания): {str(self.salary_by_cities)}")
        print(f"Доля вакансий по городам (в порядке убывания): {str(self.vacancies_share_by_cities)}")


class Vacancy:
    """Считывание файла и формирование структуры данных о нем.

        Attributes:
                dictionary (dict): название csv файла.
    """

    def __init__(self, dictionary):
        """Инициализируект объект Vacancy
               Args:
                   dictionary (dict): словарь
        >>> type(Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"})).__name__
        'Vacancy'
        >>> Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"}).name
        'Аналитик'
        >>> Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"}).area_name
        'Москва'
        >>> Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"}).salary
        55000.0
        >>> Vacancy({"name": "Аналитик", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Москва",  "published_at": "2022-07-05T18:19:30+0300"}).published_at
        2022
        """
        self.name = dictionary["name"]
        self.salary = (float(dictionary["salary_from"]) + float(dictionary["salary_to"])) / 2 * currency[
            dictionary["salary_currency"]]
        self.area_name = dictionary["area_name"]
        self.published_at = int(dictionary["published_at"][:4])


def set_border(ws, width, height):
    """Устанавливает рамки для таблицы
            Args:
                ws (openpyxl.Workbook()): Excel лист
                width (float): ширина
                height (float) высота
    """
    cell_range = f'A1:{get_column_letter(width)}{height}'
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def clear_column(ws, column):
    """Удаляет рамки для столбца таблицы
           Args:
               ws (openpyxl.Workbook()): Excel лист
               column (string): буква, соответствующая столбцу
    """
    empty = Side(border_style=None)
    for cell in ws[column]:
        cell.border = Border(top=empty, bottom=empty)


def get_vertical_chart(title, parameter_1, parameter_1_name, parameter_2,
                       parameter_2_name, labels, fig, number):
    """Создаёт вертикальную диаграмму
    """
    x = np.arange(len(labels))
    width = 0.35
    plt.rcParams['font.size'] = '8'
    ax = fig.add_subplot(number)
    ax.bar(x - width / 2, parameter_1, width, label=parameter_1_name)
    ax.bar(x + width / 2, parameter_2, width, label=parameter_2_name)
    ax.set_xticks(x, labels, rotation="vertical")
    ax.grid(axis='y')
    ax.set_title(title)
    ax.legend()


def get_horizontal_chart(title, parameter, labels, fig):
    """Создаёт вертикальную диаграмму
    """
    plt.rcParams['font.size'] = '8'
    ax = fig.add_subplot(223)
    labels = [city.replace(' ', '\n').replace('-', '-\n') for city in labels]
    y = np.arange(len(labels))
    ax.barh(y, parameter)
    ax.set_yticks(y, labels=labels, fontsize=6)
    ax.grid(axis='x')
    ax.invert_yaxis()
    ax.set_title(title)


def get_pie_chart(title, parameter, labels, fig):
    """Создаёт вертикальную диаграмму
    """
    plt.rcParams['font.size'] = '6'
    labels.insert(0, "Другие")
    parameter.insert(0, 1 - sum(parameter))
    ax = fig.add_subplot(224)
    ax.pie(parameter, labels=labels)
    ax.axis('equal')
    ax.set_title(title)
    fig.tight_layout()
    plt.savefig('graph.png')


class Report:
    """Класс создающий  png-график, excel таблицу и pdf-файл с таблицей и графиком.
       Attributes:
           dataset (DataSet): объект с данными необходимыми для построения графиков и таблиц
    """

    def __init__(self, dataset):
        """Инициализирует класс report. Структурирование данных для графиков и таблиц.
                        Args:
                            dataset (DataSet): Посчитанные данные для графиков.
        """
        self.profession = dataset.profession
        self.years_list_headers = (
            "Год", "Средняя зарплата", f"Средняя зарплата - {profession}", "Количество вакансий",
            f"Количество вакансий - {profession}")
        self.years_list_columns = [[year for year in dataset.salary_by_years],
                                   [value for value in dataset.salary_by_years.values()],
                                   [value for value in dataset.salary_by_years_for_profession.values()],
                                   [value for value in dataset.vacancies_count_by_years.values()],
                                   [value for value in dataset.vacancies_count_by_years_for_profession.values()]]

        self.cities_list_headers = ("Город", "Уровень зарплат", "", "Город", "Доля вакансий")
        self.cities_list_columns = [[city for city in dataset.salary_by_cities],
                                    [value for value in dataset.salary_by_cities.values()],
                                    ["" for i in range(len(dataset.salary_by_cities))],
                                    [city for city in dataset.vacancies_share_by_cities],
                                    [value for value in dataset.vacancies_share_by_cities.values()]]

        self.years_list_widths = [len(header) + 2 for header in self.years_list_headers]
        for i in range(len(self.years_list_columns)):
            for cell in self.years_list_columns[i]:
                self.years_list_widths[i] = max(len(str(cell)) + 2, self.years_list_widths[i])

        self.cities_list_widths = [len(header) + 2 for header in self.cities_list_headers]
        for i in range(len(self.cities_list_columns)):
            for cell in self.cities_list_columns[i]:
                self.cities_list_widths[i] = max(len(str(cell)) + 2, self.cities_list_widths[i])

    def generate_excel(self):
        """Генерирует excel файл с двумя страницами о вакансиях и таблицами в каждой:
                    по годам и по городам соотвественно
                   Returns:
                       (list): лист со станицами по годам и по городам
        """
        wb = openpyxl.Workbook()
        years_list = wb.active
        years_list.title = "Статистика по годам"
        cities_list = wb.create_sheet("Статистика по городам")
        years_list.append(self.years_list_headers)
        for cell in years_list['1']:
            cell.font = Font(bold=True)
        for i in range(len(self.years_list_columns[0])):
            years_list.append([column[i] for column in self.years_list_columns])
        cities_list.append(self.cities_list_headers)
        for cell in cities_list['1']:
            cell.font = Font(bold=True)
        for i in range(len(self.cities_list_columns[0])):
            cities_list.append([column[i] for column in self.cities_list_columns])
        for cell in cities_list['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        for i in range(1, 6):
            years_list.column_dimensions[get_column_letter(i)].width = self.years_list_widths[i - 1]
            cities_list.column_dimensions[get_column_letter(i)].width = self.cities_list_widths[i - 1]
        set_border(years_list, len(self.years_list_headers), len(self.years_list_columns[0]) + 1)
        set_border(cities_list, len(self.cities_list_headers), len(self.cities_list_columns[0]) + 1)
        clear_column('C')
        wb.save('report.xlsx')
        return [years_list, cities_list]

    def generate_image(self):
        """Метод для создания картинки с графиками.
        """
        fig = plt.figure()
        get_vertical_chart("Уровень зарплат по годам", self.years_list_columns[1], "средняя з/п",
                           self.years_list_columns[2], f"з/п {self.profession}", self.years_list_columns[0], fig,
                           221)
        get_vertical_chart("Количество вакансий по годам", self.years_list_columns[3], "Количество вакансий",
                           self.years_list_columns[4], f"Количество вакансий {self.profession}",
                           self.years_list_columns[0], fig, 222)
        get_horizontal_chart("Уровень зарплат по городам", self.cities_list_columns[1],
                             self.cities_list_columns[0], fig)
        get_pie_chart("Доля вакансий по городам", self.cities_list_columns[4], self.cities_list_columns[3], fig)
        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self):
        """Генерирует pdf файл
        """
        years_list, cities_list = self.generate_excel()
        self.generate_image()
        for cell in cities_list['E']:
            if cell.value == "Доля вакансий":
                continue
            cell.value = str(round(float(cell.value) * 100, 2)).replace('.', ',') + '%'

        env = Environment(loader=FileSystemLoader('../../Downloads'))
        pdf_template = env.get_template("pdf_template.html").render(
            {'profession': f'{self.profession}', 'image_file': "graph.png",
             'years_list': years_list, 'cities_list': cities_list})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={'enable-local-file-access': None})


currency = {"AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055, }

file_name = input("Введите название файла: ")
profession = input("Введите название профессии: ")
dataset = DataSet(file_name, profession)
dataset.print_info()
Report(dataset).generate_image()
