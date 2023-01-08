import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


class Report:
    """
    Класс для создания Excel таблицы
    """
    def __init__(self, profession, years, average_salary, average_salary_profession, count_vacancies_by_year,
                 count_vacancies_by_year_prof, file_name):
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.profession = profession
        self.file_name = file_name

    def generate_excel(self):
        df = [[self.years[i], self.average_salary[i], self.average_salary_profession[i],
               self.count_vacancies_by_year[i], self.count_vacancies_by_year_prof[i]] for i in range(len(self.years))]
        df.insert(0, ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий", f"Количество вакансий - {self.profession}"])
        df = pd.DataFrame(df, columns=None)
        with pd.ExcelWriter(self.file_name) as writer:
            df.to_excel(writer, sheet_name='Статистика по годам', index=False, header=False)
        wb = openpyxl.load_workbook(self.file_name)
        worksheet1 = wb["Статистика по годам"]
        thin = Side(border_style="thin")
        self.add_border(worksheet1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.create_max_width(worksheet1)
        wb.save(self.file_name)

    def add_border(self, worksheet, side, count_columns, rows):
        for i in range(1, count_columns):
            for row in rows:
                if i == 1:
                    worksheet[row + str(i)].alignment = Alignment(horizontal='left')
                    worksheet[row + str(i)].font = Font(bold=True)
                if worksheet[row + str(i)].internal_value is not None:
                    worksheet[row + str(i)].border = Border(top=side, bottom=side, left=side, right=side)

    def create_max_width(self, worksheet):
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value is not None:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                else:
                    dims[cell.column] = len(str(cell.value))
        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2
